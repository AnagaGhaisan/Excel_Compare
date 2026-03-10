import os
import re
import openpyxl
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font

# File paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DRAFT_TEMPLATE_PATH = os.path.join(BASE_DIR, "static/template/Draft Output.xlsx")

# Helper function to check file extensions
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def extract_no_faktur_from_description(desc: str, voucher_cat: str) -> str | None:
    if pd.isna(desc):
        return None
    
    s = str(desc).strip()
    # Memisahkan berdasarkan tanda "/"
    parts = [p.strip() for p in s.split("/")]
    
    # Logika Baru: Jika GL-JV, ambil bagian pertama (index 0)
    if str(voucher_cat).strip() == "GL-JV":
        return parts[0] if len(parts) >= 1 and parts[0] else None
    
    # Selain GL-JV, ambil bagian kedua (index 1)
    return parts[1] if len(parts) >= 2 and parts[1] else None


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalisasi nama kolom: uppercase, spasi/punctuation -> underscore, rapihin underscore.
    Contoh: 'No Voucher' -> 'NO_VOUCHER'
    """
    def clean(col):
        col = str(col).strip().upper()
        col = re.sub(r"[^A-Z0-9]+", "_", col)
        col = re.sub(r"_+", "_", col).strip("_")
        return col

    df = df.copy()
    df.columns = [clean(c) for c in df.columns]
    return df

def _parse_id_number(x):
    """
    Aman untuk angka dengan format Indonesia:
    - '77.597.727' -> 77597727
    - '7.759.773'  -> 7759773
    - '1.234,56'   -> 1234.56
    """
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)

    s = str(x).strip()
    if not s:
        return np.nan

    s = s.replace(" ", "")
    # hapus ribuan ".", ubah desimal "," jadi "."
    s = s.replace(".", "")
    s = s.replace(",", ".")
    # handle (123) -> -123
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]

    try:
        v = float(s)
        return -v if neg else v
    except:
        return np.nan

def calculate_net(row):
    debit = float(row.get("Debit Amount", 0))
    credit = float(row.get("Credit Amount", 0))
    acc_name = str(row.get("Account Name", "")).strip()

    # Pendapatan: -Debit + Credit
    if acc_name in ["Interest Bank Income", "Other Income", "Rental Income", 
                    "Repair Service Income", "Sales", "Sales Price Protection"]:
        return -debit + credit 

    # Beban: Debit - Credit
    elif acc_name in ["POP Expense", "Promotion Gift"]:
        return debit - credit
        
    # Sales Return: -Debit - Credit (Sesuai -AA10-AB10)
    elif acc_name == "Sales Return":
        return -(debit - credit) 
        
    return 0


    
def compare_files(k3_sheets: dict, coretax_sheets_1: dict, coretax_sheets_2: dict, output_dir: str) -> str:
    # k3_sheets, coretax_sheets_1, coretax_sheets_2 are already dicts of {sheet_name: DataFrame}
    # from pd.read_excel(..., sheet_name=None) in app.py — no need to re-read.

    # 1) Concatenate all K3 sheets (keep original column names for later use)
    k3 = pd.concat(k3_sheets.values(), ignore_index=True)
    print(f"K3 combined shape: {k3.shape}, columns: {list(k3.columns)}")

    # 2) BARU terapkan ekstraksi No. Faktur & Nett pada variabel 'k3'
    k3["No Faktur (key)"] = k3.apply(
        lambda row: extract_no_faktur_from_description(row.get('Description', ''), row.get('Voucher Category', '')), 
        axis=1
    )
    # Bersihkan spasi agar tidak meleset saat merge
    k3["No Faktur (key)"] = k3["No Faktur (key)"].astype(str).str.strip()
    k3["Nett"] = k3.apply(calculate_net, axis=1)

    # 2) Normalize columns for Coretax (biar NO VOUCHER / DOC_NO kebaca konsisten)
    coretax_1 = pd.concat([_normalize_columns(sheet_data) for sheet_data in coretax_sheets_1.values()], ignore_index=True)
    coretax_2 = pd.concat([_normalize_columns(sheet_data) for sheet_data in coretax_sheets_2.values()], ignore_index=True)

    # Pastikan key jadi NO_VOUCHER
    if "DOC_NO" in coretax_1.columns and "NO_VOUCHER" not in coretax_1.columns:
        coretax_1 = coretax_1.rename(columns={"DOC_NO": "NO_VOUCHER"})
    if "DOC_NO" in coretax_2.columns and "NO_VOUCHER" not in coretax_2.columns:
        coretax_2 = coretax_2.rename(columns={"DOC_NO": "NO_VOUCHER"})

    if "NO_VOUCHER" not in coretax_1.columns:
        raise ValueError("Coretax Digunggung: kolom DOC_NO / NO VOUCHER tidak ditemukan.")
    if "NO_VOUCHER" not in coretax_2.columns:
        raise ValueError("Coretax Tidak Digunggung: kolom DOC_NO / NO VOUCHER tidak ditemukan.")

    # 3) Harmonize DPP/PPN + CUSTOMER + status
    # --- Digunggung: AMOUNT_BEF_TAX = DPP, TAX_AMOUNT = PPN, CUSTOMER_NAME = CUSTOMER
    if "DPP" not in coretax_1.columns and "AMOUNT_BEF_TAX" in coretax_1.columns:
        coretax_1["DPP"] = coretax_1["AMOUNT_BEF_TAX"]
    if "PPN" not in coretax_1.columns and "TAX_AMOUNT" in coretax_1.columns:
        coretax_1["PPN"] = coretax_1["TAX_AMOUNT"]

    coretax_1["CUSTOMER"] = coretax_1["CUSTOMER_NAME"] if "CUSTOMER_NAME" in coretax_1.columns else None
    coretax_1["FP_STATUS"] = "FP Digunggung"

    # --- Tidak Digunggung: DPP = DPP, PPN = PPN, NAMA_PEMBELI = CUSTOMER
    if "DPP" not in coretax_2.columns and "AMOUNT_BEF_TAX" in coretax_2.columns:
        coretax_2["DPP"] = coretax_2["AMOUNT_BEF_TAX"]
    if "PPN" not in coretax_2.columns and "TAX_AMOUNT" in coretax_2.columns:
        coretax_2["PPN"] = coretax_2["TAX_AMOUNT"]

    if "NAMA_PEMBELI" in coretax_2.columns:
        coretax_2["CUSTOMER"] = coretax_2["NAMA_PEMBELI"]
    elif "CUSTOMER_NAME" in coretax_2.columns:
        coretax_2["CUSTOMER"] = coretax_2["CUSTOMER_NAME"]
    else:
        coretax_2["CUSTOMER"] = None

    coretax_2["FP_STATUS"] = "FP Tidak Digunggung"

    if "DEPT" in coretax_1.columns:
        coretax_1["CUSTOMER"] = coretax_1["DEPT"]
    elif "CUSTOMER_NAME" in coretax_1.columns:
        coretax_1["CUSTOMER"] = coretax_1["CUSTOMER_NAME"]
    else:
        coretax_1["CUSTOMER"] = None

    coretax_1["FP_STATUS"] = "FP Digunggung"

    # 4) Bersihin key + convert angka
    for df in (coretax_1, coretax_2):
        df["NO_VOUCHER"] = df["NO_VOUCHER"].astype(str).str.strip()
        if "DPP" in df.columns:
            df["DPP"] = df["DPP"].apply(_parse_id_number)
        else:
            df["DPP"] = 0.0
        if "PPN" in df.columns:
            df["PPN"] = df["PPN"].apply(_parse_id_number)
        else:
            df["PPN"] = 0.0

    # 5) Combine Coretax (ambil kolom penting aja)
    keep_cols_1 = [c for c in ["NO_VOUCHER", "VOUCHER_NO", "DPP", "PPN", "CUSTOMER", "FP_STATUS"] if c in coretax_1.columns]
    keep_cols_2 = [c for c in ["NO_VOUCHER", "VOUCHER_NO", "DPP", "PPN", "CUSTOMER", "FP_STATUS"] if c in coretax_2.columns]
    coretax_combined = pd.concat([coretax_1[keep_cols_1], coretax_2[keep_cols_2]], ignore_index=True)

    # 6) kalau NO_VOUCHER muncul beberapa kali, DPP/PPN dijumlah, CUSTOMER diambil first non-null, status digabung unik
    def join_unique(series):
        vals = [v for v in series.dropna().astype(str).tolist() if v.strip()]
        return "; ".join(sorted(set(vals))) if vals else None

    agg_map = {
        "DPP": "sum",
        "PPN": "sum",
        "CUSTOMER": "first",  # Takes first non-null value for CUSTOMER
        "FP_STATUS": join_unique
    }
    if "VOUCHER_NO" in coretax_combined.columns:
        agg_map["VOUCHER_NO"] = "first"

    # Aggregate the combined coretax data

    # 8) Debugging step: Check columns in Coretax_2
    print("Columns in Coretax_2:", coretax_2.columns)

    # 9) Debugging: Check if 'NO FP MODIF' exists in coretax_2
    if "NO FP MODIF" in coretax_2.columns:
        print("NO FP MODIF exists in Coretax_2.")
    else:
        print("NO FP MODIF NOT found in Coretax_2")

    # --- 1) Gabungkan Coretax Tanpa Agregasi Dahulu ---
    coretax_combined = pd.concat([coretax_1, coretax_2], ignore_index=True)

    # --- 2) Identifikasi Kasus Split yang Mirip (Jumlah Baris Sama) ---
    # Hitung berapa kali tiap voucher muncul di masing-masing file
    k3_counts = k3['No Faktur (key)'].value_counts()
    core_counts = coretax_combined['NO_VOUCHER'].value_counts()

    # Fungsi untuk menentukan apakah harus pakai row_order
    def check_split_case(v_no):
        if v_no == "-" or pd.isna(v_no): return False
        # Hanya apply jika jumlah baris di K3 dan Coretax SAMA (Misal sama-sama 2 baris)
        return k3_counts.get(v_no, 0) == core_counts.get(v_no, 0) and k3_counts.get(v_no, 0) > 1

    # Tandai baris yang masuk kategori "Kasus Mirip"
    k3['is_split_case'] = k3['No Faktur (key)'].apply(check_split_case)
    coretax_combined['is_split_case'] = coretax_combined['NO_VOUCHER'].apply(check_split_case)

    # Tambahkan row_order HANYA untuk yang is_split_case
    k3['row_order'] = 0
    k3.loc[k3['is_split_case'], 'row_order'] = k3[k3['is_split_case']].groupby('No Faktur (key)').cumcount()

    coretax_combined['row_order'] = 0
    coretax_combined.loc[coretax_combined['is_split_case'], 'row_order'] = coretax_combined[coretax_combined['is_split_case']].groupby('NO_VOUCHER').cumcount()

    # --- 3) Merge dengan Kondisi ---
    merged = pd.merge(
        k3,
        coretax_combined,
        left_on=["No Faktur (key)", "row_order"],
        right_on=["NO_VOUCHER", "row_order"],
        how="left"
    )

# 11) Compute Difference based on account type
    merged["Debit Amount"] = pd.to_numeric(merged["Debit Amount"], errors="coerce").fillna(0)
    merged["Credit Amount"] = pd.to_numeric(merged["Credit Amount"], errors="coerce").fillna(0)

    # Apply the Net calculation based on the account type
    merged["Net"] = merged.apply(calculate_net, axis=1)

    merged["DPP"] = pd.to_numeric(merged["DPP"], errors="coerce").fillna(0)
    merged["PPN"] = pd.to_numeric(merged["PPN"], errors="coerce").fillna(0)

    # Logika baru untuk menghitung Difference
    def calculate_difference(row):
        # Jika akun adalah Sales Return, gunakan rumus Debit + DPP
        if row["Account Name"] == "Sales Return":
            return row["Debit Amount"] + row["DPP"]
        # Untuk akun lainnya, gunakan rumus standar Net - DPP
        else:
            return row["Net"] - row["DPP"]

    # Terapkan fungsi ke kolom Difference
    merged["Difference"] = merged.apply(calculate_difference, axis=1)
    

    # 12) Keterangan + Customer (langsung dari kolom kanonik)
    merged["Keterangan (Digunggung/Tidak Digunngung)"] = merged["FP_STATUS"]
    merged.loc[merged["_merge"] != "both", "Keterangan (Digunggung/Tidak Digunngung)"] = "Tidak ada di Coretax"

    merged["Customer"] = merged["CUSTOMER"]
    merged.loc[merged["_merge"] != "both", "Customer"] = None

    # Debugging: Check if 'NO_FP_MODIF' exists in coretax_2
    print("Cek apakah 'NO_FP_MODIF' ada di coretax_2:", "NO_FP_MODIF" in coretax_2.columns)
    if "NO_FP_MODIF" in coretax_2.columns:
        print(coretax_2["NO_FP_MODIF"].head())  # Menampilkan beberapa nilai untuk memastikan kolom ada
    else:
        print("Kolom 'NO_FP_MODIF' tidak ditemukan di Coretax_2")

    print("Setelah merge, kolom di merged:", merged.columns)

    # Jika kolom 'NO_FP_MODIF' ada di coretax_2, tambahkan ke merged
    if "NO_FP_MODIF" in coretax_2.columns:
        merged = pd.merge(
            merged,
            coretax_2[["NO_VOUCHER", "NO_FP_MODIF"]],
            left_on="No Faktur (key)",
            right_on="NO_VOUCHER",
            how="left",
            suffixes=("", "_from_coretax2")
        )
        print("Setelah merge NO_FP_MODIF, kolom di merged:", merged.columns)
    else:
        merged["NO_FP_MODIF"] = None  # Atur sebagai None jika kolom tidak ada

    # 13) Before filling NaN, convert categorical columns to string type
    for column in merged.columns:
        if merged[column].dtype.name == 'category':  # Check if the column is categorical
            merged[column] = merged[column].astype(str)
    # Now, proceed with other operations
    merged.fillna("-", inplace=True)

    # Cek jika file Excel ada
    if not os.path.exists(DRAFT_TEMPLATE_PATH):
        raise FileNotFoundError("Draft Output.xlsx tidak ditemukan. Taruh file itu 1 folder dengan app.py")

    # Load template Excel
    wb = load_workbook(DRAFT_TEMPLATE_PATH)
    ws = wb.active

# --- 14) INISIALISASI TOTAL & VARIABEL (Sebelum Loop) ---
    debit_total = credit_total = net_total = balance_total = 0
    dpp_total = ppn_total = difference_total = 0
    
    voucher_group_mapping = {}
    current_duplicate_group = 1
    voucher_count = merged["NO_VOUCHER"].value_counts().to_dict()

    template_ws = ws 
    current_ws = ws
    current_row_in_sheet = 0 
    sheet_count = 1
    start_row = 5
    max_data_rows_per_sheet = 500000

    # Helper: copy header rows (1-4) from template to a new sheet
    def _copy_header_rows(src_ws, dst_ws, up_to_row=4):
        for row_idx in range(1, up_to_row + 1):
            for col_idx in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row_idx, col_idx)
                dst_cell = dst_ws.cell(row_idx, col_idx)
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font = src_cell.font.copy()
                    dst_cell.border = src_cell.border.copy()
                    dst_cell.fill = src_cell.fill.copy()
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment = src_cell.alignment.copy()

    # --- 15) LOOP PENULISAN (TETAP SPLIT, HASIL AKURAT) ---
# --- 15) LOOP PENULISAN ---
    for i in range(len(merged)):
        if current_row_in_sheet >= max_data_rows_per_sheet:
            sheet_count += 1
            current_ws = wb.create_sheet(title=f"Sheet{sheet_count}")
            _copy_header_rows(template_ws, current_ws)
            current_row_in_sheet = 0

        r = start_row + current_row_in_sheet
        row = merged.iloc[i]

        # 1. AMBIL NILAI DASAR (Selalu definisikan di awal loop agar aman)
        val_debit = float(row.get("Debit Amount", 0))
        val_credit = float(row.get("Credit Amount", 0))
        val_net = float(row.get("Net", 0))
        val_bal = _parse_id_number(row.get("Balance", 0))
        val_dpp = float(row.get("DPP", 0))
        val_ppn = float(row.get("PPN", 0))
        val_diff = float(row.get("Difference", 0))
        
        voucher_no = str(row.get("NO_VOUCHER", "-"))
        status = ""
        is_split = row.get('is_split_case', False)

        # 2. LOGIKA PERHITUNGAN & STATUS
        if is_split:
            # KASUS SPLIT: Tulis 1-ke-1, Total dihitung tiap baris
            status = f"Split Match ({int(row['row_order'])+1})"
            
            # Tambahkan ke Subtotal Global
            debit_total += val_debit
            credit_total += val_credit
            net_total += val_net
            balance_total += val_bal
            dpp_total += val_dpp
            ppn_total += val_ppn
            difference_total += (val_net - val_dpp)
            
            # Update nilai diff khusus tampilan split
            val_diff = val_net - val_dpp

        elif voucher_no != "-" and voucher_count.get(voucher_no, 0) > 1:
            # KASUS DUPLIKAT BIASA (KONSOLIDASI)
            if voucher_no not in voucher_group_mapping:
                voucher_group_mapping[voucher_no] = current_duplicate_group
                status = f"Duplicate {current_duplicate_group}"
                
                # Konsolidasi Sisi Kiri (GL)
                v_rows = merged[merged["NO_VOUCHER"] == voucher_no]
                val_debit = v_rows["Debit Amount"].sum()
                val_credit = v_rows["Credit Amount"].sum()
                
                # Hitung ulang Net & Diff untuk baris pertama
                combined_row = row.copy()
                combined_row["Debit Amount"] = val_debit
                combined_row["Credit Amount"] = val_credit
                val_net = calculate_net(combined_row)
                val_diff = val_net - val_dpp

                # Tambahkan ke Subtotal Global (Hanya sekali per grup)
                debit_total += val_debit
                credit_total += val_credit
                net_total += val_net
                balance_total += val_bal
                dpp_total += val_dpp
                ppn_total += val_ppn
                difference_total += val_diff
                
                current_duplicate_group += 1
            else:
                status = f"Duplicate {voucher_group_mapping[voucher_no]}"
                # Baris lanjutan di-set 0 agar tidak double counting
                val_debit = val_credit = val_net = val_dpp = val_ppn = val_diff = 0
                balance_total += val_bal # Balance tetap dihitung per baris jika perlu
        else:
            # KASUS UNIQUE
            status = "Unique"
            debit_total += val_debit
            credit_total += val_credit
            net_total += val_net
            balance_total += val_bal
            dpp_total += val_dpp
            ppn_total += val_ppn
            difference_total += val_diff

        # 3. TULIS DATA KE EXCEL (Sekarang variabel pasti terdefinisi)
        current_ws.cell(r, 1).value = row.get("Account No.")
        current_ws.cell(r, 2).value = row.get("Account Name")
        current_ws.cell(r, 3).value = row.get("Date")
        current_ws.cell(r, 4).value = row.get("Voucher Category")
        current_ws.cell(r, 5).value = row.get("Voucher No.")
        current_ws.cell(r, 6).value = row.get("Description")
        current_ws.cell(r, 7).value = val_debit
        current_ws.cell(r, 8).value = val_credit
        current_ws.cell(r, 9).value = val_net
        current_ws.cell(r, 10).value = row.get("Direction")
        current_ws.cell(r, 11).value = val_bal
        
        current_ws.cell(r, 13).value = voucher_no
        current_ws.cell(r, 14).value = row.get("NO_FP_MODIF")
        current_ws.cell(r, 15).value = val_dpp
        current_ws.cell(r, 16).value = val_ppn
        current_ws.cell(r, 17).value = val_diff
        current_ws.cell(r, 18).value = row.get("Customer")
        current_ws.cell(r, 19).value = row.get("Keterangan (Digunggung/Tidak Digunngung)")
        current_ws.cell(r, 20).value = status
        
        current_row_in_sheet += 1

    # --- 16) CETAK TOTAL KE BARIS 3 ---
    template_ws.cell(3, 7).value = debit_total
    template_ws.cell(3, 8).value = credit_total
    template_ws.cell(3, 9).value = net_total
    template_ws.cell(3, 11).value = balance_total
    template_ws.cell(3, 15).value = dpp_total
    template_ws.cell(3, 16).value = ppn_total
    template_ws.cell(3, 17).value = difference_total 

    # Penulisan Bold (Font)
    bold_font = Font(bold=True)
    for col in [7, 8, 9, 11, 15, 16, 17]:
        template_ws.cell(3, col).font = bold_font

    print(f"Selesai! Data ditulis ke {sheet_count} sheet. Total baris: {len(merged)}")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Save the workbook (preserves template formatting + all sheets)
    out_name = f"Draft_Updated_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)
    print(f"Saved output to {out_path}")

    # Dapatkan daftar nama sheet dari workbook yang baru disimpan
    sheet_names = wb.sheetnames
    
    return out_path, out_name, sheet_names


def delete_all_uploaded_files(app):
    try:
        # Cek jika direktori upload ada
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            # Hapus semua file dalam folder upload
            for filename in os.listdir(app.config['UPLOAD_FOLDER']):
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)  # Hapus file
                    print(f"Uploaded file deleted: {file_path}")
    except Exception as e:
        print(f"Error deleting uploaded files: {e}")