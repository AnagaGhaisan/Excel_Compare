import os
import pandas as pd
import openpyxl
import re
from flask import Blueprint, request, redirect, url_for, current_app
from werkzeug.utils import secure_filename
from helpers import allowed_file, delete_all_uploaded_files
import uuid

recap_bp = Blueprint("recap_bp", __name__)

# Gunakan huruf besar secara konsisten untuk variabel global
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "static/template/Ekualisasi_Darft_Output.xlsx")


def calculate_row_nett(row):
    acc_name = str(row.get("Account Name", "")).strip()
    try:
        debit = float(row.get("Debit Amount", 0))
        credit = float(row.get("Credit Amount", 0))
    except:
        debit, credit = 0.0, 0.0

    # Logika: -Debit + Credit
    if acc_name in [
        "Interest Bank Income",
        "Other Income",
        "Other Operating Income",
        "Rental Income",
        "Repair Service Income",
        "Sales",
        "Sales Price Protection",
    ]:
        return -debit + credit  # <--- Disamakan agar konsisten

    elif acc_name in ["POP Expense", "Promotion Gift"]:
        return debit - credit

    elif acc_name == "Sales Return":
        return -(debit - credit)

    return 0


def process_recap_2_files(source_path, ppn_path, output_dir):
    # Cek menggunakan variabel global TEMPLATE_PATH
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template tidak ditemukan di path: {TEMPLATE_PATH}")

    try:
        # --- 1. AMBIL DATA DARI REKAP PPN ---
        df_ppn_raw = pd.read_excel(ppn_path, header=None)
        ppn_monthly_values = {}
        for i, row in df_ppn_raw.iterrows():
            row_vals = [str(x).strip().lower() for x in row.values]
            if "jumlah penyerahan" in row_vals:
                for m in range(1, 13):
                    val = row.iloc[5 + m]
                    ppn_monthly_values[m] = float(val) if pd.notnull(val) else 0
                break

        # --- 2. HITUNG DATA DARI DRAFT GL ---
        xls_gl = pd.ExcelFile(source_path)
        all_gl_dfs = []
        for sheet in xls_gl.sheet_names:
            df_raw = pd.read_excel(xls_gl, sheet_name=sheet, header=None)
            h_row = -1
            for i, row in df_raw.iterrows():
                if "Account Name" in [str(x) for x in row.values]:
                    h_row = i
                    break
            if h_row != -1:
                df = df_raw.iloc[h_row + 1 :].copy()
                df.columns = [str(x).strip() for x in df_raw.iloc[h_row]]
                df = df.iloc[:, ~df.columns.duplicated()]
                all_gl_dfs.append(df)

        df_combined = pd.concat(all_gl_dfs, ignore_index=True)
        df_combined["Date"] = pd.to_datetime(
            df_combined["Date"], dayfirst=True, errors="coerce"
        )
        df_combined["Month"] = df_combined["Date"].dt.month
        df_combined["Calculated_Nett"] = df_combined.apply(calculate_row_nett, axis=1)

        summary_month = (
            df_combined.groupby(["Account Name", "Month"])["Calculated_Nett"]
            .sum()
            .to_dict()
        )
        summary_total_acc = (
            df_combined.groupby("Account Name")["Calculated_Nett"].sum().to_dict()
        )
        summary_all_month_total = (
            df_combined.groupby("Month")["Calculated_Nett"].sum().to_dict()
        )

        # --- 3. ISI TEMPLATE (Perbaikan Variabel di sini) ---
        wb = openpyxl.load_workbook(TEMPLATE_PATH)  # Menggunakan TEMPLATE_PATH
        ws = wb.active

        # Update Bagian Atas & Total AH20
        total_gl_year = 0
        for r in range(8, 20):
            acc_cell = ws.cell(row=r, column=3)
            if acc_cell.value:
                acc_name = str(acc_cell.value).strip()
                val = summary_total_acc.get(acc_name, 0)
                ws.cell(row=r, column=5).value = val
                ws.cell(row=r, column=5).number_format = "#,##0"
                total_gl_year += val
        ws.cell(row=20, column=34).value = total_gl_year

        # Grid Bulanan & Baris TOTAL (Baris 39)
        header_row_gl = 23
        month_start_row = 25
        total_row_idx = 39
        account_to_col = {}

        for c in range(1, ws.max_column + 1):
            h_val = ws.cell(row=header_row_gl, column=c).value
            if h_val and "cfm." in str(h_val):
                match = re.search(r"-\s*(.+)$", str(h_val))
                if match:
                    account_to_col[match.group(1).strip()] = c + 1

        grand_ppn = 0
        grand_gl_total = 0
        grand_selisih = 0
        acc_column_totals = {acc: 0 for acc in account_to_col.keys()}

        for m in range(1, 13):
            t_row = month_start_row + (m - 1)
            m_ppn = ppn_monthly_values.get(m, 0)
            ws.cell(row=t_row, column=3).value = m_ppn
            ws.cell(row=t_row, column=3).number_format = "#,##0"
            grand_ppn += m_ppn

            for acc, t_col in account_to_col.items():
                val = summary_month.get((acc, m), 0)
                ws.cell(row=t_row, column=t_col).value = val
                ws.cell(row=t_row, column=t_col).number_format = "#,##0"
                acc_column_totals[acc] += val

            m_gl_total = summary_all_month_total.get(m, 0)
            ws.cell(row=t_row, column=32).value = m_gl_total
            ws.cell(row=t_row, column=32).number_format = "#,##0"
            grand_gl_total += m_gl_total

            m_selisih = m_ppn - m_gl_total
            ws.cell(row=t_row, column=34).value = m_selisih
            ws.cell(row=t_row, column=34).number_format = "#,##0"
            grand_selisih += m_selisih
            if m_selisih != 0:
                ws.cell(row=t_row, column=34).font = openpyxl.styles.Font(
                    color="FF0000", bold=True
                )

        # ISI BARIS TOTAL (BARIS 39)
        ws.cell(row=total_row_idx, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=total_row_idx, column=3).value = grand_ppn
        ws.cell(row=total_row_idx, column=3).number_format = "#,##0"

        for acc, t_col in account_to_col.items():
            ws.cell(row=total_row_idx, column=t_col).value = acc_column_totals[acc]
            ws.cell(row=total_row_idx, column=t_col).number_format = "#,##0"

        ws.cell(row=total_row_idx, column=32).value = grand_gl_total
        ws.cell(row=total_row_idx, column=32).number_format = "#,##0"
        ws.cell(row=total_row_idx, column=34).value = grand_selisih
        ws.cell(row=total_row_idx, column=34).number_format = "#,##0"

        for c in [3, 32, 34] + list(account_to_col.values()):
            ws.cell(row=total_row_idx, column=c).font = openpyxl.styles.Font(bold=True)

        out_name = f"Final_Ekualisasi_{os.path.basename(source_path)}"
        out_path = os.path.join(output_dir, out_name)
        wb.save(out_path)

        return out_name, wb.sheetnames

    except Exception as e:
        raise ValueError(f"Proses Gagal: {str(e)}")


# ... (Route upload_recap tetap sama seperti sebelumnya) ...
@recap_bp.route("/upload_recap", methods=["POST"])
def upload_recap():
    # 1. Pengecekan keberadaan file berdasarkan name di HTML
    if "k3_file" not in request.files or "ppn_file" not in request.files:
        return "Kesalahan: File Draft (GL) dan Rekap PPN wajib diupload", 400

    source_file = request.files["k3_file"]
    ppn_file = request.files["ppn_file"]

    if source_file.filename == "" or ppn_file.filename == "":
        return "Kesalahan: Nama file tidak boleh kosong", 400

    if source_file and ppn_file:
        # 2. Definisikan path
        unique_id = str(uuid.uuid4())[:8]

        # 3. Tambahkan unique_id ke nama file
        s_name = f"{unique_id}_{secure_filename(source_file.filename)}"
        p_name = f"{unique_id}_{secure_filename(ppn_file.filename)}"

        # 4. Gabungkan dengan path folder upload
        s_path = os.path.join(current_app.config["UPLOAD_FOLDER"], s_name)
        p_path = os.path.join(current_app.config["UPLOAD_FOLDER"], p_name)

        # 5. Simpan file fisik
        source_file.save(s_path)
        ppn_file.save(p_path)

        # Menggunakan folder khusus recap agar tidak campur dengan compare
        output_dir = current_app.config["OUTPUT_RECAP_FOLDER"]
        os.makedirs(output_dir, exist_ok=True)

        try:
            # 4. Jalankan proses
            out_name, sheets = process_recap_2_files(s_path, p_path, output_dir)

            # --- BAGIAN PENGHAPUSAN DIHAPUS ---
            delete_all_uploaded_files(current_app)
            # Dengan menghapus baris di atas, file tetap tersimpan permanen.

            # 5. Redirect ke hasil dengan mode 'recap'
            return redirect(
                url_for(
                    "show_comparison",
                    updated_file=out_name,
                    mode="recap",
                    sheets=",".join(sheets),
                )
            )
        except Exception as e:
            return f"Terjadi kesalahan saat memproses: {str(e)}", 500

    return "Format file tidak valid", 400